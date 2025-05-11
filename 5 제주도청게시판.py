from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
import pandas as pd
import os
import openpyxl
from openpyxl.styles import Border, Side, Alignment

initial_page_url = 'https://www.jeju.go.kr/news/news/law/jeju2.htm'
output_folder = '.'
excel_file_name = '제주도 기관 공고.xlsx'
excel_output_path = os.path.join(output_folder, excel_file_name)

service = ChromeService(ChromeDriverManager().install())
options = Options()
options.add_argument('--headless')
options.add_argument('--disable-gpu')
options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')
options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')

driver = webdriver.Chrome(service=service, options=options)
data = []

try:
    driver.get(initial_page_url)
    time.sleep(2)
    for page in range(1, 11):  # 1~10페이지
        print(f"{page}페이지 데이터 추출 중...")
        table_element = driver.find_element(By.CSS_SELECTOR, '#gosiBody')
        rows = table_element.find_elements(By.TAG_NAME, 'tr')
        for row in rows:
            cols = row.find_elements(By.TAG_NAME, 'td')
            if cols:
                row_data = [col.text.strip() for col in cols]
                # 3번째 컬럼(a 태그의 href 추출)
                href = ""
                try:
                    a_tag = cols[2].find_element(By.TAG_NAME, "a")
                    href = a_tag.get_attribute("href")
                except Exception:
                    href = ""
                row_data.append(href)  # 컬럼7로 추가
                data.append(row_data)
        # 다음 페이지로 이동
        try:
            next_btn = driver.find_element(By.XPATH, f"//a[text()='{page+1}']")
            driver.execute_script("arguments[0].scrollIntoView(true);", next_btn)
            time.sleep(1)
            next_btn.click()
            time.sleep(2)
        except:
            print(f"총 {page}페이지의 데이터를 추출했습니다.")
            break
except Exception as e:
    print(f'데이터 추출 중 오류: {e}')

driver.quit()

if data:
    # 엑셀 파일이 이미 있으면 기존 파일을 불러옴
    if os.path.exists(excel_output_path):
        wb = openpyxl.load_workbook(excel_output_path)
        # 기존에 '제주도청' 시트가 있으면 해당 시트 사용
        if '제주도청' in wb.sheetnames:
            ws = wb['제주도청']
            # 기존 데이터만 삭제 (B5셀부터)
            for row in range(5, ws.max_row + 1):
                for col in range(2, ws.max_column + 1):
                    ws.cell(row=row, column=col).value = None
        else:
            ws = wb.create_sheet('제주도청')
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '제주도청'

    # B5셀부터 데이터만 입력 (row=5, column=2)
    for i, row in enumerate(data):
        for j, value in enumerate(row):
            ws.cell(row=5+i, column=2+j, value=value)
        # 행 높이를 30으로 설정
        ws.row_dimensions[5+i].height = 30

    # 테두리 적용
    for row in range(5, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    wb.save(excel_output_path)
    print(f"데이터 업데이트 완료: {excel_output_path}")
else:
    print("추출된 데이터가 없습니다.")