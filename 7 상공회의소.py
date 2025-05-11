from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from datetime import datetime
import time
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import os
import re

# 설정 변수
CONFIG = {
    'URL': 'https://jejucci.korcham.net/front/board/boardContentsListPage.do?boardId=10088&menuId=379',
    'EXCEL_FILE_NAME': '제주도 기관 공고.xlsx',
    'SHEET_NAME': '상공회의소',
    'MAX_PAGES': 3,  # 수집할 최대 페이지 수
    'START_CELL': 'B5',  # 시작 셀 위치
    'ROW_HEIGHT': 25,  # 행 높이
}

def get_column_number(column_letter):
    """엑셀 열 문자를 숫자로 변환 (A=1, B=2, ...)"""
    result = 0
    for i, char in enumerate(reversed(column_letter.upper())):
        result += (ord(char) - ord('A') + 1) * (26 ** i)
    return result

def parse_cell_reference(cell_reference):
    """셀 참조(예: 'B5')를 행과 열 번호로 분리"""
    match = re.match(r'([A-Za-z]+)(\d+)', cell_reference)
    if not match:
        raise ValueError(f"Invalid cell reference: {cell_reference}")
    column_letter, row = match.groups()
    return int(row), get_column_number(column_letter)

# CSS 선택자 설정
SELECTORS = {
    'BOARD_LIST': '#contentsList > div.contents_detail > div.boardlist > table > tbody > tr',
    'NUMBER': 'td.c_number',
    'TITLE': 'td.title.c_title',
    'DATE': 'td.c_reg_dt',
    'PAGINATION': '#contentsList > div.contents_detail > div.paging.paging_area > span.mobileoff_in > a'
}

def setup_driver():
    """Selenium 웹드라이버 설정"""
    options = Options()
    options.add_argument('--headless')  # 헤드리스 모드
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--window-size=1920,1080')
    
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)

def is_notice(number_text):
    """공지사항 여부 확인"""
    try:
        int(number_text)
        return False
    except ValueError:
        return True

def get_max_page(driver):
    """실제 마지막 페이지 번호 확인"""
    try:
        # 페이지 번호 요소들 가져오기
        page_elements = driver.find_elements(By.CSS_SELECTOR, SELECTORS['PAGINATION'])
        if not page_elements:
            return 1
        
        # 페이지 번호 추출
        page_numbers = []
        for element in page_elements:
            try:
                page_num = int(element.text.strip())
                page_numbers.append(page_num)
            except ValueError:
                continue
        
        return max(page_numbers) if page_numbers else 1
    except Exception as e:
        print(f"페이지 번호 확인 중 오류 발생: {str(e)}")
        return 1

def update_excel_with_data(df):
    """기존 Excel 파일의 서식을 유지하면서 데이터만 업데이트"""
    try:
        # 테두리 스타일 설정
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # START_CELL에서 행과 열 번호 추출
        start_row, start_col = parse_cell_reference(CONFIG['START_CELL'])
        
        # 기존 파일이 있는 경우
        if os.path.exists(CONFIG['EXCEL_FILE_NAME']):
            wb = load_workbook(CONFIG['EXCEL_FILE_NAME'])
            
            # 시트가 없는 경우 새로 생성
            if CONFIG['SHEET_NAME'] not in wb.sheetnames:
                ws = wb.create_sheet(CONFIG['SHEET_NAME'])
            else:
                ws = wb[CONFIG['SHEET_NAME']]
            
            # 데이터프레임의 데이터를 기존 셀에 업데이트
            for i, row in df.iterrows():
                for j, value in enumerate(row):
                    cell = ws.cell(row=start_row + i, column=start_col + j)
                    cell.value = value
                    cell.border = thin_border  # 테두리 적용
                    
                # 행 높이 설정
                ws.row_dimensions[start_row + i].height = CONFIG['ROW_HEIGHT']
            
            # 변경사항 저장
            wb.save(CONFIG['EXCEL_FILE_NAME'])
            
        else:
            # 새로운 파일 생성
            with pd.ExcelWriter(CONFIG['EXCEL_FILE_NAME'], engine='openpyxl') as writer:
                # 시작 행에서 1을 빼는 이유: pandas의 startrow는 0-based index 사용
                df.to_excel(writer, sheet_name=CONFIG['SHEET_NAME'], 
                          index=False, startrow=start_row-1, startcol=start_col-1)
                
                wb = writer.book
                ws = wb[CONFIG['SHEET_NAME']]
                
                # 데이터가 있는 셀에 테두리 적용
                for i in range(len(df)):
                    for j in range(len(df.columns)):
                        cell = ws.cell(row=start_row + i, column=start_col + j)
                        cell.border = thin_border
                    # 행 높이 설정
                    ws.row_dimensions[start_row + i].height = CONFIG['ROW_HEIGHT']
                
                # 파일 저장
                wb.save(CONFIG['EXCEL_FILE_NAME'])
        
        print("Excel 파일 업데이트 완료")
        
    except Exception as e:
        print(f"Excel 업데이트 중 오류 발생: {str(e)}")

def extract_data():
    """웹사이트에서 데이터 추출"""
    driver = setup_driver()
    all_data = []
    
    try:
        driver.get(CONFIG['URL'])
        time.sleep(2)  # 초기 페이지 로딩 대기
        
        # 실제 마지막 페이지 확인
        max_page = min(get_max_page(driver), CONFIG['MAX_PAGES'])
        print(f"전체 수집 페이지 수: {max_page}")
        
        current_page = 1
        while current_page <= max_page:
            print(f"페이지 {current_page} 처리 중...")
            
            # 게시글 목록 대기
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, SELECTORS['BOARD_LIST']))
            )
            
            # 페이지 로딩을 위한 짧은 대기
            time.sleep(2)
            
            # 게시글 목록 추출
            posts = driver.find_elements(By.CSS_SELECTOR, SELECTORS['BOARD_LIST'])
            
            for post in posts:
                try:
                    # 번호
                    number = post.find_element(By.CSS_SELECTOR, SELECTORS['NUMBER']).text.strip()
                    
                    # 공지사항 여부 확인
                    notice_flag = is_notice(number)
                    
                    # 첫 페이지가 아니고 공지사항이면 건너뛰기
                    if current_page > 1 and notice_flag:
                        continue
                    
                    # 공지사항인 경우 번호를 "<공지>"로 변경
                    if notice_flag:
                        number = "<공지>"
                    
                    # 제목
                    title = post.find_element(By.CSS_SELECTOR, SELECTORS['TITLE']).text.strip()
                    
                    # 작성일
                    date = post.find_element(By.CSS_SELECTOR, SELECTORS['DATE']).text.strip()
                    
                    all_data.append({
                        '번호': number,
                        '제목': title,
                        '작성일': date
                    })
                    
                except Exception as e:
                    print(f"게시글 처리 중 오류 발생: {str(e)}")
                    continue
            
            # 다음 페이지로 이동
            if current_page < max_page:
                try:
                    # JavaScript로 페이지 이동
                    next_page_num = current_page + 1
                    driver.execute_script(f"go_Page({next_page_num})")
                    time.sleep(2)  # 페이지 로딩 대기
                    current_page += 1
                except Exception as e:
                    print(f"페이지 이동 중 오류 발생: {str(e)}")
                    break
            else:
                break
        
        # 데이터를 DataFrame으로 변환
        df = pd.DataFrame(all_data)
        
        # 기존 Excel 파일의 서식을 유지하면서 데이터만 업데이트
        update_excel_with_data(df)
        
        print(f"데이터가 {CONFIG['EXCEL_FILE_NAME']} 파일로 저장되었습니다.")
        
    except Exception as e:
        print(f"데이터 추출 중 오류 발생: {str(e)}")
    
    finally:
        driver.quit()

if __name__ == "__main__":
    extract_data() 