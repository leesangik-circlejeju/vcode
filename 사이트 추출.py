import threading
import tkinter as tk
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
import os
import openpyxl
from openpyxl.styles import Border, Side, Alignment
import re
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

SETTINGS_FILE = "설정.xlsx"
SETTINGS_SHEET = "설정"

def make_template(str1, str2):
    prefix = ''
    for a, b in zip(str1, str2):
        if a == b:
            prefix += a
        else:
            break
    rev1, rev2 = str1[::-1], str2[::-1]
    suffix = ''
    for a, b in zip(rev1, rev2):
        if a == b:
            suffix = a + suffix
        else:
            break
    template = prefix + '{page}' + suffix
    return template

def extract_table_root_xpath(xpath):
    return re.sub(r'/tr\[.*?\]/td\[.*?\]$', '', xpath)

def extract_table_root_selector(selector):
    return selector.split('>')[0].strip()

def is_url_template(url1, url2):
    return url1 != url2 and '{page}' in make_template(url1, url2)

def is_xpath_template(xpath1, xpath2):
    return xpath1 != xpath2 and '{page}' in make_template(xpath1, xpath2)

def is_selector_template(sel1, sel2):
    return sel1 != sel2 and '{page}' in make_template(sel1, sel2)

def load_sites_from_excel(settings_path, sheet_name="설정"):
    wb = openpyxl.load_workbook(settings_path, data_only=True)
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    sites = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # 사이트명 없으면 skip
            continue
        site_settings = {headers[i]: row[i] for i in range(len(headers))}
        sites.append(site_settings)
    return sites

def crawl_site(site, progress_callback, stop_flag):
    # 엑셀 한 행의 모든 값을 변수로 할당
    URL_PAGE1 = site.get("URL_PAGE1")
    SELECTOR_PAGE1 = site.get("SELECTOR_PAGE1")
    XPATH_PAGE1 = site.get("XPATH_PAGE1")
    URL_PAGE2 = site.get("URL_PAGE2")
    SELECTOR_PAGE2 = site.get("SELECTOR_PAGE2")
    XPATH_PAGE2 = site.get("XPATH_PAGE2")
    TABLE_XPATH = site.get("TABLE_XPATH")
    TABLE_SELECTOR = site.get("TABLE_SELECTOR")
    EXCEL_FILE_NAME = site.get("EXCEL_FILE_NAME")
    EXCEL_SHEET_NAME = site.get("EXCEL_SHEET_NAME")
    PAGE_COUNT = int(site.get("PAGE_COUNT", 5))
    CELL_FROM = site.get("CELL_FROM")
    excel_output_path = os.path.join('.', EXCEL_FILE_NAME)

    # 1. 사이트명 표시
    progress_callback(f"[진행] 사이트명: {site.get('사이트명')}")

    url_template = make_template(URL_PAGE1, URL_PAGE2)
    selector_template = make_template(SELECTOR_PAGE1, SELECTOR_PAGE2)
    xpath_template = make_template(XPATH_PAGE1, XPATH_PAGE2)

    service = ChromeService(ChromeDriverManager().install())
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    driver = webdriver.Chrome(service=service, options=options)
    data = []
    try:
        for page in range(1, PAGE_COUNT + 1):
            if stop_flag['stop']:
                progress_callback(f"[{site.get('사이트명')}] 중지되었습니다.")
                break
            # 2. 페이지 처리중 표시
            progress_callback(f"[{site.get('사이트명')}] {page}페이지 처리중...")
            # 1. URL이 다르면 URL 템플릿 우선
            if is_url_template(URL_PAGE1, URL_PAGE2):
                url = url_template.format(page=page)
                driver.get(url)
                time.sleep(2)
            # 2. XPATH가 다르면 XPATH 템플릿으로 페이지 버튼 클릭
            elif is_xpath_template(XPATH_PAGE1, XPATH_PAGE2):
                if page == 1:
                    driver.get(URL_PAGE1)
                    time.sleep(2)
                else:
                    xpath = xpath_template.format(page=page)
                    try:
                        page_btn = driver.find_element(By.XPATH, xpath)
                        driver.execute_script("arguments[0].scrollIntoView(true);", page_btn)
                        page_btn.click()
                        time.sleep(2)
                    except Exception as e:
                        progress_callback(f"[{site.get('사이트명')}] {page}페이지 XPATH 버튼 클릭 실패: {e}")
                        break
            # 3. SELECTOR가 다르면 SELECTOR 템플릿으로 페이지 버튼 클릭
            elif is_selector_template(SELECTOR_PAGE1, SELECTOR_PAGE2):
                if page == 1:
                    driver.get(URL_PAGE1)
                    time.sleep(2)
                else:
                    selector = selector_template.format(page=page)
                    try:
                        page_btn = driver.find_element(By.CSS_SELECTOR, selector)
                        driver.execute_script("arguments[0].scrollIntoView(true);", page_btn)
                        page_btn.click()
                        time.sleep(2)
                    except Exception as e:
                        progress_callback(f"[{site.get('사이트명')}] {page}페이지 SELECTOR 버튼 클릭 실패: {e}")
                        break
            else:
                progress_callback(f"[{site.get('사이트명')}] 페이지 전환 규칙을 찾을 수 없습니다.")
                break

            # 테이블 루트 자동 추출
            if TABLE_XPATH:
                table_root = extract_table_root_xpath(TABLE_XPATH)
                table_element = driver.find_element(By.XPATH, table_root)
            else:
                table_root = extract_table_root_selector(TABLE_SELECTOR)
                table_element = driver.find_element(By.CSS_SELECTOR, table_root)

            rows = table_element.find_elements(By.TAG_NAME, 'tr')
            for row in rows:
                cols = row.find_elements(By.TAG_NAME, 'td')
                if cols:
                    row_data = [col.text.strip() for col in cols]
                    link = ""
                    for col in cols:
                        try:
                            a_tag = col.find_element(By.TAG_NAME, "a")
                            link = a_tag.get_attribute("href")
                            break
                        except Exception:
                            continue
                    row_data.append(link)
                    data.append(row_data)
        else:
            # 3. 데이터 추출 완료 표시
            progress_callback(f"[{site.get('사이트명')}] 데이터 추출 완료.")
    except Exception as e:
        progress_callback(f'[{site.get("사이트명")}] 데이터 추출 중 오류: {e}')
    finally:
        driver.quit()

    # 엑셀 저장 (기존 코드와 동일)
    if data:
        if os.path.exists(excel_output_path):
            wb = openpyxl.load_workbook(excel_output_path)
            if EXCEL_SHEET_NAME in wb.sheetnames:
                ws = wb[EXCEL_SHEET_NAME]
                for row in range(5, ws.max_row + 1):
                    for col in range(2, ws.max_column + 1):
                        ws.cell(row=row, column=col).value = None
            else:
                ws = wb.create_sheet(EXCEL_SHEET_NAME)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = EXCEL_SHEET_NAME

        col_letter, start_row = coordinate_from_string(CELL_FROM)
        start_col = column_index_from_string(col_letter)

        for i, row in enumerate(data):
            for j, value in enumerate(row):
                ws.cell(row=start_row + i, column=start_col + j, value=value)
            ws.row_dimensions[start_row + i].height = 30

        for row in range(start_row, ws.max_row + 1):
            for col in range(start_col, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

        wb.save(excel_output_path)
        # 4. 완료되었습니다 표시
        progress_callback(f"[{site.get('사이트명')}] 완료되었습니다.")
    else:
        progress_callback(f"[{site.get('사이트명')}] 추출된 데이터가 없습니다.")

# --- GUI 구성 ---
class CrawlerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("사이트별 고시공고 추출기")
        self.root.geometry("420x260")
        self.stop_flag = {'stop': False}
        self.thread = None

        self.label = tk.Label(root, text="엑셀 설정파일에 등록된 모든 사이트를 순차 추출합니다.", font=("맑은 고딕", 12))
        self.label.pack(pady=20, anchor="center")

        self.progress = tk.Label(root, text="", font=("맑은 고딕", 11), justify="center")
        self.progress.pack(pady=10, anchor="center")

        btn_frame = tk.Frame(root)
        btn_frame.pack(pady=10, anchor="center")

        self.start_btn = tk.Button(btn_frame, text="시작", width=10, command=self.start_crawling)
        self.start_btn.pack(side=tk.LEFT, padx=5)

        self.stop_btn = tk.Button(btn_frame, text="중지", width=10, command=self.stop_crawling, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=5)

        self.exit_btn = tk.Button(btn_frame, text="종료", width=10, command=self.exit_app)
        self.exit_btn.pack(side=tk.LEFT, padx=5)

    def update_progress(self, msg):
        self.progress.config(text=msg)
        self.root.update_idletasks()

    def start_crawling(self):
        self.stop_flag['stop'] = False
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.progress.config(text="크롤링 시작...")
        self.thread = threading.Thread(target=self.crawl_all_sites)
        self.thread.start()

    def crawl_all_sites(self):
        sites = load_sites_from_excel(SETTINGS_FILE, SETTINGS_SHEET)
        for site in sites:
            if self.stop_flag['stop']:
                self.update_progress("전체 작업이 중지되었습니다.")
                break
            crawl_site(site, self.update_progress, self.stop_flag)
        else:
            self.update_progress("모든 사이트 추출이 완료되었습니다.")

        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)

    def stop_crawling(self):
        self.stop_flag['stop'] = True
        self.stop_btn.config(state=tk.DISABLED)
        self.start_btn.config(state=tk.NORMAL)

    def exit_app(self):
        if self.thread and self.thread.is_alive():
            self.stop_flag['stop'] = True
            self.thread.join()
        self.root.destroy()

def create_sample_settings_excel(path, sheet_name="설정"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = [
        "사이트명", "URL_PAGE1","SELECTOR_PAGE1","XPATH_PAGE1","URL_PAGE2","SELECTOR_PAGE2",
        "XPATH_PAGE2", "TABLE_SELECTOR","TABLE_XPATH",  "EXCEL_FILE_NAME", "EXCEL_SHEET_NAME", "PAGE_COUNT", "CELL_FROM"
    ]
    ws.append(headers)
    ws.append([
        "제주도청",
        "https://www.jeju.go.kr/news/news/law/jeju2.htm",
        "#app > div.paging-container > div > ul > li:nth-child(1) > a",
        '//*[@id="app"]/div[4]/div/ul/li[1]/a',
        "https://www.jeju.go.kr/news/news/law/jeju2.htm",
        "#app > div.paging-container > div > ul > li:nth-child(2) > a",
        '//*[@id="app"]/div[4]/div/ul/li[2]/a',
        '#gosiBody > tr:nth-child(1) > td.no',
        '//*[@id="gosiBody"]/tr[1]/td[1]',
        '제주도 기관 공고.xlsx',
        '제주도청',
        5,
        'B5'
    ])
  
    wb.save(path)

if __name__ == "__main__":
    # 필요할 때만 샘플 설정파일 생성
    if not os.path.exists(SETTINGS_FILE):
        create_sample_settings_excel(SETTINGS_FILE, SETTINGS_SHEET)

    root = tk.Tk()
    app = CrawlerGUI(root)
    root.mainloop()