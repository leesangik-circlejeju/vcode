import requests
from lxml import html
import openpyxl
from openpyxl.utils import column_index_from_string
from datetime import datetime
from urllib.parse import urljoin
import time  # time 모듈 임포트
from openpyxl.styles import Alignment

# =============================================================================
# --- 전역 실행 설정 변수 ---
TARGET_URL_BASE = "https://www.jje.go.kr/board/list.jje?boardId=BBS_0000507&listRow=20&listCel=1&menuCd=DOM_000000103003009000&orderBy=REGISTER_DATE%20DESC&paging=ok&startPage="
MAX_PAGES_TO_CRAWL = 5
TARGET_YEAR_TO_EXTRACT = "all"

# --- 엑셀 파일 저장 설정 변수 ---
OUTPUT_FILENAME = "제주도 기관 공고.xlsx"
TARGET_SHEET_NAME = "도교육청"
START_CELL_FOR_DATA = "B6"
# "링크" 항목을 마지막 열로 이동
EXCEL_HEADER = ["번호", "제목", "학교", "작성일", "접수마감", "조회수", "링크"]
# =============================================================================


def crawl_jiles_notices_lxml(url_base, max_pages, target_year):
    """
    제주특별자치도 도교육청 공지사항 목록에서 특정 년도의 자료를 추출합니다.
    lxml과 XPath를 사용하며, 접수마감일이 오늘 이후인 공고만 추출합니다.
    """
    # --- 크롤링할 요소의 XPath 표현식을 변수로 정의 ---
    TABLE_XPATH = "//div[@id='sub_contentnw']/div[2]/table"
    ITEM_ROW_XPATH = "./tbody/tr"
    NUMBER_XPATH = ".//td[1]"
    CATEGORY_XPATH = ".//td[2]"
    TITLE_LINK_XPATH = ".//td[@class='title']/a"
    ATTACHMENT_XPATH = ".//td[3]"
    DATE_XPATH = ".//td[4]"
    FINDATE_XPATH = ".//td[5]"
    HIT_XPATH = ".//td[6]"

    results = []
    today = datetime.now().date()
    date_formats_to_try = ["%Y.%m.%d", "%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M"]

    for page in range(1, max_pages + 1):
        target_url = url_base + str(page)
        print(f"크롤링 중... URL: {target_url}")
        try:
            response = requests.get(target_url, timeout=10)  # 타임아웃 설정 (초 단위)
            response.raise_for_status()
            tree = html.fromstring(response.content.decode('utf-8'))

            table = tree.xpath(TABLE_XPATH)
            if not table:
                print("테이블을 찾을 수 없습니다.")
                break
            table = table[0]
            rows = table.xpath(ITEM_ROW_XPATH)

            for row in rows:
                try:
                    number_element = row.xpath(NUMBER_XPATH)[0] if row.xpath(NUMBER_XPATH) else None
                    number_str = number_element.text.strip() if number_element is not None else ""

                    title_element = row.xpath(TITLE_LINK_XPATH)[0] if row.xpath(TITLE_LINK_XPATH) else None
                    title = title_element.text.strip() if title_element is not None else ""
                    link = urljoin(target_url, title_element.get('href')) if title_element is not None and title_element.get('href') else ""

                    school_element = row.xpath(ATTACHMENT_XPATH)[0] if row.xpath(ATTACHMENT_XPATH) else None
                    school = school_element.text.strip() if school_element is not None else ""

                    date_element = row.xpath(DATE_XPATH)[0] if row.xpath(DATE_XPATH) else None
                    date_str = date_element.text.strip() if date_element is not None else ""

                    findate_element = row.xpath(FINDATE_XPATH)[0] if row.xpath(FINDATE_XPATH) else None
                    findate_str = findate_element.text.strip() if findate_element is not None else ""

                    hit_element = row.xpath(HIT_XPATH)[0] if row.xpath(HIT_XPATH) else None
                    hit_str = hit_element.text.strip() if hit_element is not None else ""

                    date = None
                    for fmt in date_formats_to_try:
                        try:
                            date = datetime.strptime(date_str, fmt).date()
                            break
                        except ValueError:
                            continue

                    findate = None
                    for fmt in date_formats_to_try:
                        try:
                            findate = datetime.strptime(findate_str, fmt).date()
                            break
                        except ValueError:
                            continue

                    if date and findate:
                        if target_year == "all" or date.year == int(target_year) if target_year != "all" else True:
                            if findate >= today:
                                # 데이터 추가 순서를 변경된 헤더 순서에 맞춤
                                results.append([number_str, title, school, date_str, findate_str, hit_str, link])
                    else:
                        print(f"날짜 형식 인식 오류: 작성일={date_str}, 접수마감일={findate_str}")

                except IndexError as e:
                    print(f"데이터 추출 오류: {e}, 해당 행을 건너뜁니다.")
                    continue

            time.sleep(1)
        except requests.exceptions.RequestException as e:
            print(f"요청 오류: {e}")
            continue
        except requests.exceptions.SSLError as e:
            print(f"SSL 오류: {e}")
            continue
        except Exception as e:
            print(f"기타 오류: {e}")
            break
        except IndexError:
            print("더 이상 페이지가 없습니다.")
            break

    return results


def save_to_excel(filename, sheet_name, start_cell, header, data):
    """
    크롤링한 데이터를 엑셀 파일에 저장하며, 숫자 값은 숫자로 저장하고
    기존 시트를 유지하며 지정된 시트에만 데이터를 업로드하고, 서식을 적용합니다.
    """
    try:
        workbook = openpyxl.load_workbook(filename)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

    start_row = int(start_cell[1:])
    start_col = column_index_from_string(start_cell[0])

    # 헤더 쓰기 및 서식 적용 (시트가 새로 생성된 경우에만)
    if sheet.max_row < start_row:
        header_alignment = Alignment(horizontal='center', vertical='top')
        for col_num, header_text in enumerate(header, start=start_col):
            cell = sheet.cell(row=start_row - 1, column=col_num, value=header_text)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = header_alignment

    # 데이터 쓰기 및 서식 적용
    data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    for row_num, row_data in enumerate(data, start=start_row):
        for col_num, cell_data in enumerate(row_data, start=start_col):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.alignment = data_alignment
            # 변경된 헤더 순서에 맞춰 숫자 컬럼 인덱스 조정
            if col_num - start_col + 1 in [1, 6]: # "번호"와 "조회수" 컬럼
                try:
                    cell.value = int(cell_data)
                except ValueError:
                    cell.value = cell_data
            else:
                cell.value = cell_data

    # 컬럼 너비 자동 조정 (데이터를 모두 쓴 후에)
    for col in sheet.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(filename)
    print(f"'{filename}' 파일의 '{sheet_name}' 시트에 저장 완료.")


if __name__ == "__main__":
    crawled_data = crawl_jiles_notices_lxml(TARGET_URL_BASE, MAX_PAGES_TO_CRAWL, TARGET_YEAR_TO_EXTRACT)
    if crawled_data:
        save_to_excel(OUTPUT_FILENAME, TARGET_SHEET_NAME, START_CELL_FOR_DATA, EXCEL_HEADER, crawled_data)
    else:
        print("오늘 이후의 접수 마감일인 공고가 없습니다.")