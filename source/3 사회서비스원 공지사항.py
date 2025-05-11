import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Border, Side, Alignment

# =============================================================================
# --- 전역 실행 설정 변수 ---
# 크롤링 대상 URL의 기본 경로
TARGET_URL_BASE = "https://jeju.pass.or.kr/notice-board/index"
# 크롤링할 최대 페이지 수
MAX_PAGES_TO_CRAWL = 5
# 추출할 공지사항의 년도 ("all" 또는 "YYYY" 형식)
TARGET_YEAR_TO_EXTRACT = "all" # 예: "2023", "2024", "all"

# --- 엑셀 파일 저장 설정 변수 ---
# 저장할 엑셀 파일 이름
OUTPUT_FILENAME = "제주도 기관 공고.xlsx"
# 데이터를 저장할 시트 이름
TARGET_SHEET_NAME = "사회서비스원"
# 데이터가 시작될 엑셀 셀 (헤더는 이 셀의 한 행 위)
START_CELL_FOR_DATA = "B6"
# 엑셀 파일의 헤더 목록 (이 내용이 START_CELL_FOR_DATA의 한 행 위에 저장됩니다)
EXCEL_HEADER = ["번호", "제목", "링크", "작성자", "조회수", "첨부파일", "등록일"]
# =============================================================================


def crawl_jeju_pass_notices_year_filter(url_base, max_pages, target_year):
    """
    제주특별자치도 사회서비스원 공지사항 목록에서 특정 년도의 자료만 추출합니다.
    td 클래스 이름들을 변수로 관리합니다.
    """
    # --- 크롤링할 요소의 CSS 클래스 이름을 변수로 정의 ---
    NOTICE_ITEM_CLASS = "notice" # 공지사항 각 항목(행)의 클래스
    NUMBER_CLASS = "W8 board-cell-number" # "번호" 컬럼의 클래스
    TITLE_CLASS = "title board-cell-subject" # "제목" 및 "링크" 컬럼 (a 태그 포함)의 클래스
    ATTACHMENT_TD_CLASS = "W10 m-br file board-cell-file" # "첨부파일" 컬럼 (td 태그)의 클래스
    ATTACHMENT_SPAN_CLASS = "sr-only" # 첨부파일 여부를 나타내는 span 태그의 클래스 (첨부파일 TD 내부에 있음)
    WRITER_CLASS = "W10 nowrap m-br board-cell-writer" # "작성자" 컬럼의 클래스
    HIT_CLASS = "W10 nowrap m-br board-cell-hit" # "조회수" 컬럼의 클래스
    DATE_CLASS = "W13 m-br board-cell-date" # "등록일" 컬럼의 클래스
    # ------------------------------------------------------

    all_data = []
    for page in range(1, max_pages + 1):
        print(f"페이지 {page} 크롤링 중...")
        current_url = f"{url_base}/page/{page}"
        try:
            response = requests.get(current_url)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')

            notice_list_container = soup.find('div', class_="tstyle_list board-list-table")
            if not notice_list_container:
                print(f"페이지 {page}: 공지사항 목록 컨테이너를 찾을 수 없습니다.")
                continue

            notice_items = notice_list_container.find_all('tr', class_=NOTICE_ITEM_CLASS)
            # 고정 공지사항 (tr class가 notice가 아닌 경우)도 포함하려면 find_all('tr') 후 클래스로 분기해야 할 수 있습니다.
            # 현재는 tr.notice만 찾도록 되어 있습니다.
            if not notice_items:
                print(f"페이지 {page}: 공지사항 항목(tr.{NOTICE_ITEM_CLASS})을 찾을 수 없습니다.")
                # 페이지에 항목이 없는 경우, 다음 페이지로 이동
                continue

            # 일반 게시글 항목 (tr class가 없는 경우 등)도 고려해야 한다면 find_all('tr') 후 처리 로직 추가 필요
            # 현재는 notice_items에 고정 공지사항만 담길 수 있습니다.
            # 일반 게시글은 tr 태그에 class가 없는 경우가 많으므로, 해당 부분도 크롤링하려면 find_all('tr')로 가져온 후
            # 각 tr을 순회하며 notice 클래스가 있는지 확인하는 방식이 더 견고할 수 있습니다.

            # 예시: 모든 tr 태그를 가져와서 처리하는 방식으로 변경 (일반 게시글 포함 시)
            all_tr_items = notice_list_container.find_all('tr')

            for item in all_tr_items: # 모든 tr을 순회
                # 고정 공지사항이 아닌 일반 게시글인지 확인 (클래스가 notice가 아니거나, 번호가 숫자인 경우)
                # 웹사이트 구조에 따라 일반 게시글 tr의 특정 클래스를 확인하는 것이 더 정확할 수 있습니다.
                # 현재는 notice 클래스가 없는 tr을 일반 게시글로 간주하고 처리 로직을 진행합니다.

                number_element = item.find('td', class_=NUMBER_CLASS)
                number_text = number_element.get_text(strip=True) if number_element else ""
                # 번호가 '공지' 또는 빈 문자열이 아닌 경우에만 숫자로 변환 시도
                number = int(number_text) if number_text.isdigit() else number_text

                title_element = item.find('td', class_=TITLE_CLASS)
                title_link_element = title_element.find('a') if title_element else None
                title = title_link_element.get_text(strip=True) if title_link_element else ""
                link = "https://jeju.pass.or.kr" + title_link_element['href'] if title_link_element and 'href' in title_link_element.attrs else ""

                attachment_td = item.find('td', class_=ATTACHMENT_TD_CLASS)
                attachment_element = attachment_td.find('span', class_=ATTACHMENT_SPAN_CLASS) if attachment_td else None
                attachment = "○" if attachment_element and attachment_element.get_text(strip=True) == "첨부파일 있음" else ""

                writer_element = item.find('td', class_=WRITER_CLASS)
                writer = writer_element.get_text(strip=True) if writer_element else ""

                hit_element = item.find('td', class_=HIT_CLASS)
                hit_text = hit_element.get_text(strip=True) if hit_element else ""
                hit = int(hit_text) if hit_text.isdigit() else hit_text

                date_element = item.find('td', class_=DATE_CLASS)
                date = date_element.get_text(strip=True) if date_element else ""

                # --- 디버깅 출력 ---
                print(f"  - 추출 데이터: 번호='{number}', 제목='{title}', 날짜='{date}'")
                # -------------

                # 목표 년도 필터링 및 데이터 유효성 검사
                # 번호가 있고 제목이 비어있지 않은 경우에만 데이터 추가
                if (target_year == "all" or date.startswith(target_year)):
                    if number and title: # 번호와 제목 모두 유효해야 추가
                        all_data.append([number, title, link, writer, hit, attachment, date])
                    else:
                         # 번호나 제목이 누락되어 스킵되는 항목을 명확히 출력
                         print(f"  - 스킵됨 (번호 또는 제목 누락): 번호='{number}', 제목='{title}', 날짜='{date}'")


        except requests.exceptions.RequestException as e:
            print(f"페이지 {page} 크롤링 실패: {e}")
            continue # 오류 발생 시 해당 페이지 건너뛰고 다음 페이지 시도
        except Exception as e:
            print(f"페이지 {page} 처리 중 오류 발생: {e}")
            continue # 예외 발생 시에도 다음 페이지 처리를 시도하도록 함


    # 추출된 데이터가 번호 순으로 정렬되지 않은 경우 정렬 (옵션)
    # 보통 웹사이트 목록은 최신글 순이지만, 필요에 따라 번호 순으로 다시 정렬할 수 있습니다.
    # 단, '공지' 등 숫자가 아닌 번호가 있을 경우 오류 발생 가능. 숫자형 번호만 정렬 대상으로 하거나 예외 처리 필요.
    # try:
    #     all_data.sort(key=lambda x: int(x[0]) if isinstance(x[0], (int, str)) and str(x[0]).isdigit() else float('inf'))
    # except Exception as e:
    #     print(f"데이터 정렬 중 오류 발생: {e}")

    return all_data


def update_excel(data, filename, sheet_name, start_cell):
    """기존 엑셀 파일의 지정된 시트에 데이터를 업데이트하고 테두리, 자동 줄바꿈, 정렬을 적용합니다."""
    try:
        workbook = openpyxl.load_workbook(filename)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            start_col = openpyxl.utils.column_index_from_string(start_cell[0])
            start_row = int(start_cell[1:])

            # 데이터 삭제 (헤더 아래 부분만)
            # 마지막 데이터가 쓰여진 행 이후부터 삭제
            max_row_to_delete_from = start_row
            # 현재 시트에 데이터가 얼마나 있는지 확인하여 정확한 삭제 범위 설정
            if sheet.max_row >= start_row:
                 # start_row 부터 sheet.max_row 까지 삭제
                 sheet.delete_rows(start_row, sheet.max_row - start_row + 1)


            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            left_top_alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
            center_alignment = Alignment(horizontal='center', vertical='top')

            # 새로운 데이터 추가 및 스타일 적용
            for row_num, row_data in enumerate(data):
                # 데이터 행은 start_row에서 시작
                for col_num, cell_value in enumerate(row_data):
                    # 데이터 열은 start_col에서 시작
                    cell = sheet.cell(row=start_row + row_num, column=start_col + col_num, value=cell_value)
                    cell.border = thin_border
                    # 컬럼 인덱스 (0부터 시작)는 EXCEL_HEADER의 순서를 따릅니다.
                    # EXCEL_HEADER = ["번호", "제목", "링크", "작성자", "조회수", "첨부파일", "등록일"]
                    if col_num in [4, 5, 6]: # 조회수, 첨부파일, 등록일 컬럼 인덱스
                        cell.alignment = center_alignment
                    else: # 번호, 제목, 링크, 작성자
                        cell.alignment = left_top_alignment

                    if col_num == 0 or col_num == 4: # 번호, 조회수 컬럼 인덱스
                        if isinstance(cell_value, int):
                            cell.number_format = '#,##0'
                        elif isinstance(cell_value, str) and str(cell_value).isdigit(): # 문자열 형태의 숫자도 고려
                             cell.number_format = '#,##0'
                        else:
                             cell.number_format = '@' # 숫자가 아닌 경우 텍스트 형식으로

            # 헤더 스타일 적용 (헤더는 start_row - 1 에 위치)
            header_row = start_row - 1 # 데이터 시작 행(B6)의 한 행 위는 B5
            header_col_start = start_col # 데이터 시작 열(B)과 같은 열

            # EXCEL_HEADER 변수 사용
            for col_offset, header_text in enumerate(EXCEL_HEADER):
                # 헤더 셀 위치: 행은 start_row - 1 (B5), 열은 start_col (B)부터 시작
                header_cell = sheet.cell(row=header_row, column=header_col_start + col_offset, value=header_text) # 헤더 값도 다시 쓰기
                header_cell.border = thin_border
                # 컬럼 인덱스 (0부터 시작)는 EXCEL_HEADER의 순서를 따릅니다.
                if col_offset in [4, 5, 6]: # 조회수, 첨부파일, 등록일 컬럼 인덱스
                    header_cell.alignment = center_alignment
                else:
                    header_cell.alignment = left_top_alignment


            workbook.save(filename)
            print(f"'{filename}' 파일의 '{sheet_name}' 시트에 데이터가 업데이트되고 테두리, 자동 줄바꿈, 정렬, 숫자 형식이 적용되었습니다.")

        else:
            print(f"오류: '{filename}' 파일에 '{sheet_name}' 시트가 존재하지 않습니다. 새로 생성합니다.")
            save_to_excel_with_style(data, filename, sheet_name, start_cell)

    except FileNotFoundError:
        print(f"오류: '{filename}' 파일을 찾을 수 없습니다. 새로 생성합니다.")
        save_to_excel_with_style(data, filename, sheet_name, start_cell)
    except Exception as e:
        print(f"엑셀 파일 업데이트 실패: {e}")


def save_to_excel_with_style(data, filename, sheet_name, start_cell):
    """추출된 공지사항 데이터를 지정된 엑셀 파일, 시트명, 시작 셀로 저장하고 스타일을 적용합니다."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    # EXCEL_HEADER 변수 사용
    header = EXCEL_HEADER
    start_col = openpyxl.utils.column_index_from_string(start_cell[0])
    start_row = int(start_cell[1:])
     # EXCEL_HEADER 변수 사용
    for col_num, header_text in enumerate(header):
        # 헤더 셀 위치: 행은 start_row - 1 (B5), 열은 start_col (B)부터 시작
        sheet.cell(row=start_row - 1, column=start_col + col_num, value=header_text)

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    left_top_alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='top')

    for row_num, row_data in enumerate(data):
        for col_num, cell_value in enumerate(row_data):
            cell = sheet.cell(row=start_row + row_num, column=start_col + col_num, value=cell_value)
            cell.border = thin_border
            # 컬럼 인덱스 (0부터 시작)는 EXCEL_HEADER의 순서를 따릅니다.
            # EXCEL_HEADER = ["번호", "제목", "링크", "작성자", "조회수", "첨부파일", "등록일"]
            if col_num in [4, 5, 6]: # 조회수, 첨부파일, 등록일 컬럼 인덱스
                cell.alignment = center_alignment
            else:
                cell.alignment = left_top_alignment
            if col_num == 0 or col_num == 4: # 번호, 조회수 컬럼 인덱스
                 if isinstance(cell_value, int):
                    cell.number_format = '#,##0'
                 elif isinstance(cell_value, str) and str(cell_value).isdigit(): # 문자열 형태의 숫자도 고려
                    cell.number_format = '#,##0'
                 else:
                    cell.number_format = '@' # 숫자가 아닌 경우 텍스트 형식으로


    # EXCEL_HEADER 변수 사용
    for col_offset, header_text in enumerate(header):
        # 헤더 셀 위치: 행은 start_row - 1 (B5), 열은 start_col (B)부터 시작
        header_cell = sheet.cell(row=start_row - 1, column=start_col + col_offset)
        header_cell.border = thin_border
        # 컬럼 인덱스 (0부터 시작)는 EXCEL_HEADER의 순서를 따릅니다.
        if col_offset in [4, 5, 6]: # 조회수, 첨부파일, 등록일 컬럼 인덱스
            header_cell.alignment = center_alignment
        else:
            header_cell.alignment = left_top_alignment

    try:
        workbook.save(filename)
        print(f"데이터가 '{filename}' 파일의 '{sheet_name}' 시트 '{start_cell}' 셀부터 저장되었고 스타일이 적용되었습니다.")
    except Exception as e:
        print(f"엑셀 파일 저장 실패: {e}")


if __name__ == "__main__":
    print(f"{TARGET_URL_BASE} 에서 공지사항 크롤링 시작 (최대 {MAX_PAGES_TO_CRAWL} 페이지, 년도: {TARGET_YEAR_TO_EXTRACT})...")
    notice_data = crawl_jeju_pass_notices_year_filter(
        TARGET_URL_BASE,
        MAX_PAGES_TO_CRAWL,
        TARGET_YEAR_TO_EXTRACT
    )

    if notice_data:
        print(f"총 {len(notice_data)} 건의 데이터를 추출했습니다. 엑셀 파일에 저장합니다.")
        update_excel(
            notice_data,
            filename=OUTPUT_FILENAME,
            sheet_name=TARGET_SHEET_NAME,
            start_cell=START_CELL_FOR_DATA
        )
    else:
        print("크롤링된 데이터가 없습니다. 엑셀 파일 업데이트를 건너뜁니다.")