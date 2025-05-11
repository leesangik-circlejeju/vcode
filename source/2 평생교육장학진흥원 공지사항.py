import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Border, Side, Alignment
import time # 시간 모듈 임포트
# URL 파싱 및 구성 도구 임포트
from urllib.parse import urlparse, parse_qs, urlencode, urlunparse, urljoin

# =============================================================================
# --- 전역 실행 설정 변수 ---
# 크롤링 대상 URL의 기본 경로 (첫 페이지 URL 포함)
TARGET_URL_BASE = "https://jiles.or.kr/community/notice/notice.htm?page=1&category=181"
# 크롤링할 최대 페이지 수
MAX_PAGES_TO_CRAWL = 3
# 추출할 공지사항의 년도 ("all" 또는 "YYYY" 형식)
TARGET_YEAR_TO_EXTRACT = "all" # 예: "2023", "2024", "all"

# --- 엑셀 파일 저장 설정 변수 ---
# 저장할 엑셀 파일 이름
OUTPUT_FILENAME = "제주도 기관 공고.xlsx"
# 데이터를 저장할 시트 이름
TARGET_SHEET_NAME = "평생교육장학진흥원"
# 데이터가 시작될 엑셀 셀 (헤더는 이 셀의 한 행 위)
START_CELL_FOR_DATA = "B6"
# 엑셀 파일의 헤더 목록 (이 내용이 START_CELL_FOR_DATA의 한 행 위에 저장됩니다)
EXCEL_HEADER = ["번호", "분류", "제목", "링크", "첨부", "등록일", "조회수"]
# =============================================================================


# 함수 이름 변경: 대상 사이트와 라이브러리 반영
def crawl_jiles_notices_bs4(url_base, max_pages, target_year):
    """
    제주특별자치도 평생교육장학진흥원 공지사항 목록에서 특정 년도의 자료를 추출합니다.
    BeautifulSoup와 CSS 선택자를 사용합니다.
    """
    # --- 크롤링할 요소의 CSS 선택자를 변수로 정의 ---
    # 공지사항 목록 테이블 (이 테이블 내에서 tr을 찾습니다.)
    TABLE_SELECTOR = "table.table-list"
    # 각 게시글 항목 TR (테이블 기준 상대 경로)
    ITEM_ROW_SELECTOR = "tbody tr"

    # 각 컬럼 TD 및 그 안의 데이터 선택자
    NUMBER_SELECTOR = ".no" # 번호 셀
    CATEGORY_SELECTOR = ".category" # 분류 셀
    TITLE_LINK_SELECTOR = ".title a" # 제목 텍스트와 링크
    # 첨부파일 이미지 선택자 (alt="파일첨부" 속성을 가진 img 태그)
    attachment_selector = ".attach img[alt='파일첨부']"
    # 작성자 관련 선택자 삭제됨
    DATE_SELECTOR = ".wdate" # 등록일 셀
    # 조회수 셀 선택자 변경: .hit -> .read
    HIT_SELECTOR = ".read" # 조회수 셀
    # ----------------------------------------------------------------------

    all_data = []

    # 링크 구성을 위한 기본 URL 결정
    parsed_base_url = urlparse(url_base)
    link_base = urlunparse((parsed_base_url.scheme, parsed_base_url.netloc, parsed_base_url.path.rsplit('/', 1)[0] + '/', '', '', ''))


    for page in range(1, max_pages + 1):
        print(f"페이지 {page} 크롤링 시작 중...")
        # 현재 페이지 URL 구성
        parsed_url = urlparse(url_base)
        query_params = parse_qs(parsed_url.query)
        query_params['page'] = [str(page)]
        updated_query = urlencode(query_params, doseq=True)
        current_url = urlunparse((parsed_url.scheme, parsed_url.netloc, parsed_url.path, parsed_url.params, updated_query, parsed_url.fragment))


        print(f"페이지 {page}: URL 접속 시도 - {current_url}")
        try:
            response = requests.get(current_url)
            print(f"페이지 {page}: 응답 수신 완료.")
            response.raise_for_status()
            print(f"페이지 {page}: HTTP 상태코드 정상 ({response.status_code}).")

            # BeautifulSoup으로 HTML 파싱
            soup = BeautifulSoup(response.text, 'html.parser')
            print(f"페이지 {page}: HTML 파싱 완료.")

            # 테이블 컨테이너 찾기
            list_table = soup.select_one(TABLE_SELECTOR)
            if not list_table:
                print(f"페이지 {page}: 오류 - 공지사항 목록 테이블({TABLE_SELECTOR})을 찾을 수 없습니다.")
                continue


            # 게시글 항목 TR 모두 찾기
            all_tr_items = list_table.select(ITEM_ROW_SELECTOR)

            if not all_tr_items:
                 print(f"페이지 {page}: 게시글 항목(tr)을 찾을 수 없습니다.")
                 continue

            print(f"페이지 {page}: 총 {len(all_tr_items)}개의 tr 항목 찾음. 데이터 추출 시작.")

            for item in all_tr_items: # 각 tr 순회
                # CSS 선택자를 사용하여 데이터 추출

                # 번호 추출
                number_element = item.select_one(NUMBER_SELECTOR)
                number_text = number_element.get_text(strip=True) if number_element else ""
                number = int(number_text) if number_text.isdigit() else number_text

                # 분류 추출
                category_element = item.select_one(CATEGORY_SELECTOR)
                extracted_category = category_element.get_text(strip=True) if category_element else ""
                category = "모집공고" if extracted_category else ""


                # 제목 및 링크 추출
                title_link_element = item.select_one(TITLE_LINK_SELECTOR)
                title = title_link_element.get_text(strip=True) if title_link_element else ""
                link_suffix = title_link_element['href'] if title_link_element and 'href' in title_link_element.attrs else ""
                link = urljoin(link_base, link_suffix) if link_suffix else ""


                # 첨부파일 유무 확인 (alt="파일첨부"인 img 태그의 존재 여부)
                attachment_img_element = item.select_one(attachment_selector)
                attachment = "○" if attachment_img_element else ""


                # 등록일 추출
                date_element = item.select_one(DATE_SELECTOR)
                date = date_element.get_text(strip=True) if date_element else ""

                # 조회수 추출 (변경된 HIT_SELECTOR 사용)
                hit_element = item.select_one(HIT_SELECTOR)
                hit_text = hit_element.get_text(strip=True) if hit_element else ""
                # 조회수가 숫자가 아닌 경우 (예: '조회수')를 대비하여 isdigit() 체크
                hit = int(hit_text) if hit_text and hit_text.isdigit() else hit_text


                # 목표 년도 필터링 및 데이터 유효성 검사
                # EXCEL_HEADER 순서: ["번호", "분류", "제목", "링크", "첨부", "등록일", "조회수"]
                if (target_year == "all" or (date and date.startswith(target_year))):
                    if (isinstance(number, int) or number == '공지') and title:
                         all_data.append([number, category, title, link, attachment, date, hit])


            print(f"페이지 {page}: 데이터 추출 및 필터링 완료.")


        except requests.exceptions.RequestException as e:
            print(f"페이지 {page} 크롤링 실패 (Request 오류): {e}")
            continue
        except Exception as e:
            print(f"페이지 {page} 처리 중 예상치 못한 일반 오류 발생: {e}")
            # import traceback
            # traceback.print_exc()
            continue

    return all_data


# 엑셀 업데이트 및 저장 함수 (기존 시트 유지 기능 포함)
def update_excel(data, filename, sheet_name, start_cell):
    """
    엑셀 파일의 지정된 시트를 업데이트하거나 새로 생성하고
    데이터, 테두리, 자동 줄바꿈, 정렬, 컬럼 너비를 적용합니다.
    기존 파일의 다른 시트는 유지됩니다.
    """
    try:
        print(f"엑셀 파일 '{filename}' 처리 시작.")
        try:
            workbook = openpyxl.load_workbook(filename)
            print(f"기존 파일 '{filename}' 로드 성공.")
            file_exists = True
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            print(f"'{filename}' 파일이 존재하지 않아 새로 생성합니다.")
            file_exists = False

        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"'{sheet_name}' 시트 찾음. 데이터 업데이트를 시작합니다 (데이터 {len(data)}건).")
            start_row_index = int(start_cell[1:])
            if sheet.max_row >= start_row_index:
                 print(f"기존 데이터 삭제: 행 {start_row_index}부터 {sheet.max_row}까지")
                 sheet.delete_rows(start_row_index, sheet.max_row - start_row_index + 1)

        else:
            print(f"'{sheet_name}' 시트가 존재하지 않아 새로 생성합니다 (데이터 {len(data)}건).")
            if not file_exists and 'Sheet' in workbook.sheetnames and len(workbook.sheetnames) == 1:
                 print("새 워크북의 기본 시트('Sheet')를 삭제합니다.")
                 workbook.remove(workbook['Sheet'])

            sheet = workbook.create_sheet(sheet_name)


        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        left_top_alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
        center_aligned_cols_indices = [4, 5, 6] # 첨부, 등록일, 조회수
        center_alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

        start_col_letter = start_cell[0]
        start_col_index = openpyxl.utils.column_index_from_string(start_col_letter)
        start_row_index = int(start_cell[1:])


        header_row = start_row_index - 1
        header_col_start_index = start_col_index
        print("헤더 작성 및 스타일 적용...")
        for col_offset, header_text in enumerate(EXCEL_HEADER):
             header_cell = sheet.cell(row=header_row, column=header_col_start_index + col_offset, value=header_text)
             header_cell.border = thin_border
             if col_offset in center_aligned_cols_indices:
                 header_cell.alignment = center_alignment
             else:
                 header_cell.alignment = left_top_alignment
        print("헤더 작성 및 스타일 적용 완료.")


        if data:
            print("새 데이터 추가 및 스타일 적용...")
            for row_num, row_data in enumerate(data):
                for col_num, cell_value in enumerate(row_data):
                    cell = sheet.cell(row=start_row_index + row_num, column=start_col_index + col_num, value=cell_value)
                    cell.border = thin_border
                    if col_num in center_aligned_cols_indices:
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = left_top_alignment

                number_col_index = 0
                hit_col_index = 6
                if number_col_index < len(row_data):
                    cell_number = sheet.cell(row=start_row_index + row_num, column=start_col_index + number_col_index)
                    if isinstance(row_data[number_col_index], int):
                        cell_number.number_format = '#,##0'
                    else:
                         cell_number.number_format = '@'

                if hit_col_index < len(row_data):
                     cell_hit = sheet.cell(row=start_row_index + row_num, column=start_col_index + hit_col_index)
                     if isinstance(row_data[hit_col_index], int):
                        cell_hit.number_format = '#,##0'
                     elif isinstance(row_data[hit_col_index], str) and str(row_data[hit_col_index]).isdigit():
                         cell_hit.number_format = '#,##0'
                     else:
                         cell_hit.number_format = '@'


            print("새 데이터 추가 완료.")
        else:
            print("저장할 데이터가 없습니다. 데이터 추가 단계를 건너뜹니다.")


        print("컬럼 너비 조절 적용...")
        col_widths = {
            0: 8,   # 번호 (B열)
            1: 12,  # 분류 (C열)
            2: 40,  # 제목 (D열)
            3: 50,  # 링크 (E열)
            4: 8,   # 첨부 (F열)
            5: 12,  # 등록일 (G열)
            6: 10   # 조회수 (H열)
        }
        for col_index, width in col_widths.items():
            actual_col_index = start_col_index + col_index
            col_letter = openpyxl.utils.get_column_letter(actual_col_index)
            sheet.column_dimensions[col_letter].width = width
        print("컬럼 너비 조절 적용 완료.")


        try:
            workbook.save(filename)
            print(f"'{filename}' 파일이 성공적으로 업데이트 또는 저장되었습니다.")
        except Exception as e:
            print(f"엑셀 파일 저장 실패: {e}")
            # import traceback
            # traceback.print_exc()

    except Exception as e:
        print(f"엑셀 파일 처리 중 예상치 못한 오류 발생: {e}")
        # import traceback
        # traceback.print_exc()


if __name__ == "__main__":
    print("사회서비스원 공지사항을 수집합니다.")
    print("-" * 30)

    print(f"대상 URL 기본 경로: {TARGET_URL_BASE}")
    print(f"최대 크롤링 페이지: {MAX_PAGES_TO_CRAWL}")
    print(f"대상 년도: {TARGET_YEAR_TO_EXTRACT}")
    print(f"출력 파일: {OUTPUT_FILENAME}")
    print(f"대상 시트: {TARGET_SHEET_NAME}")
    print(f"데이터 시작 셀: {START_CELL_FOR_DATA} (헤더는 {int(START_CELL_FOR_DATA[1:])-1}행 시작)")
    print("-" * 30)

    notice_data = crawl_jiles_notices_bs4(
        TARGET_URL_BASE,
        MAX_PAGES_TO_CRAWL,
        TARGET_YEAR_TO_EXTRACT
    )

    print("-" * 30)
    if notice_data:
        print(f"크롤링 및 필터링된 총 데이터 수: {len(notice_data)} 건")
        print("엑셀 파일 처리를 시작합니다 (기존 시트 유지).")
        update_excel(
            notice_data,
            filename=OUTPUT_FILENAME,
            sheet_name=TARGET_SHEET_NAME,
            start_cell=START_CELL_FOR_DATA
        )
    else:
        print("크롤링되거나 필터링된 데이터가 없습니다. 엑셀 파일 업데이트를 건너뛰므로 기존 파일 내용이 유지됩니다.")

    print("작업이 완료되었습니다")
    time.sleep(3)