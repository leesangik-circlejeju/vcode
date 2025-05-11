# 설정 변수
CONFIG = {
    'URL': 'https://hrdc.hrdkorea.or.kr/hrdc/jeju',
    'EXCEL_FILE_NAME': '제주도 기관 공고.xlsx',  # 엑셀 파일
    'SHEET_NAME': '산업인력',
    'MAX_PAGES': 3,  # 수집할 최대 페이지 수
    'START_CELL': 'B5',  # 시작 셀
    'ROW_HEIGHT': 25,    # 행 높이
    'HEADLESS': True,
}
COLUMNS = ["번호", "제목", "작성자", "작성일", "조회수", "링크"]

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from webdriver_manager.chrome import ChromeDriverManager
import time
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

def setup_driver():
    options = Options()
    options.add_argument('--headless')  # headless 모드 활성화 (브라우저 창이 뜨지 않음)
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--window-size=1920,1080')
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)

def go_to_jeju_notice_page(driver):
    try:
        driver.get(CONFIG['URL'])
        time.sleep(2)
        notice_btn = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#community > a"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", notice_btn)
        time.sleep(1)
        try:
            notice_btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", notice_btn)
        time.sleep(2)
        current_url = driver.current_url
        print("이동 후 URL:", current_url)
        if "notice" in current_url or "알림마당" in driver.page_source:
            print("알림마당 하위 사이트로 정상 이동했습니다.")
            return True
        else:
            print("알림마당 하위 사이트 이동 실패 또는 예상과 다른 페이지입니다.")
            return False
    except Exception as e:
        print(f"알림마당 이동 중 오류: {e}")
        return False

def extract_notice_table(driver):
    # 표의 모든 행(tr) 추출
    rows = driver.find_elements(By.CSS_SELECTOR, "body > div > section > form > div > table > tbody > tr")
    data = []
    for row in rows:
        tds = row.find_elements(By.TAG_NAME, "td")
        if len(tds) < 6:
            continue
        번호 = tds[0].text.strip()
        # 제목에 링크가 있을 경우 링크 텍스트와 onclick 속성 추출
        try:
            제목_요소 = tds[1].find_element(By.TAG_NAME, "a")
            제목 = 제목_요소.text.strip()
            onclick = 제목_요소.get_attribute("onclick")
            # onclick에서 숫자만 추출
            링크_번호 = onclick.split("'")[1] if onclick else ""
            링크 = f"https://hrdc.hrdkorea.or.kr/hrdc/{링크_번호}"
        except:
            제목 = tds[1].text.strip()
            링크 = ""
        작성자 = tds[2].text.strip()
        작성일 = tds[3].text.strip()
        추천수 = tds[4].text.strip()
        조회수 = tds[5].text.strip()
        data.append({
            "번호": 번호,
            "제목": 제목,
            "작성자": 작성자,
            "작성일": 작성일,
            "추천수": 추천수,
            "조회수": 조회수,
            "링크": 링크
        })
    return data

def save_notice_to_excel(data, filename, sheet_name="나라일터", start_cell="B5"):
    df = pd.DataFrame(data, columns=COLUMNS)
    startrow = int(''.join(filter(str.isdigit, start_cell))) - 1
    startcol = ord(''.join(filter(str.isalpha, start_cell)).upper()) - ord('A')

    # 기존 파일이 있으면 해당 시트만 업데이트, 없으면 새로 생성
    if os.path.exists(filename):
        wb = load_workbook(filename)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]
        # 기존 데이터 영역 지우기
        for row in ws.iter_rows(min_row=startrow+1, max_row=ws.max_row, min_col=startcol+1, max_col=startcol+len(COLUMNS)):
            for cell in row:
                cell.value = None
                cell.border = None
        # 데이터 입력 및 서식 적용
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
        for i, row in enumerate(df.values):
            for j, value in enumerate(row):
                cell = ws.cell(row=startrow+1+i, column=startcol+1+j, value=value)
                cell.border = border
            ws.row_dimensions[startrow+1+i].height = CONFIG['ROW_HEIGHT']
        wb.save(filename)
    else:
        # 새 파일 생성
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow, startcol=startcol)
        # 테두리 및 행높이 적용
        wb = load_workbook(filename)
        ws = wb[sheet_name]
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
        for i, row in enumerate(df.values):
            for j, value in enumerate(row):
                cell = ws.cell(row=startrow+1+i, column=startcol+1+j)
                cell.border = border
            ws.row_dimensions[startrow+1+i].height = CONFIG['ROW_HEIGHT']
        wb.save(filename)

def go_to_notice_page(driver, page_num):
    if page_num == 1:
        return True
    try:
        page_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, f"body > div > section > div:nth-child(6) > div > a[onclick*='fnMove({page_num})']"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", page_btn)
        time.sleep(0.5)
        page_btn.click()
        time.sleep(2)
        return True
    except Exception as e:
        print(f"{page_num}페이지 이동 중 오류: {e}")
        return False

# 메인 실행부
if __name__ == "__main__":
    driver = setup_driver()
    if go_to_jeju_notice_page(driver):
        all_data = []
        for page in range(1, CONFIG['MAX_PAGES'] + 1):
            if page > 1:
                if not go_to_notice_page(driver, page):
                    break
            time.sleep(1)
            all_data.extend(extract_notice_table(driver))
        save_notice_to_excel(
            all_data,
            CONFIG['EXCEL_FILE_NAME'],
            sheet_name=CONFIG.get('SHEET_NAME', '알림마당'),
            start_cell=CONFIG.get('START_CELL', 'B5')
        )
        print(f"엑셀 파일로 저장 완료: {CONFIG['EXCEL_FILE_NAME']}")
    # driver.quit() 