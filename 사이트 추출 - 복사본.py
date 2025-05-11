import threading
import tkinter as tk
from tkinter import messagebox
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
import os
import openpyxl
from openpyxl.styles import Border, Side, Alignment
import re
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

# =======================
# 1, 2페이지의 정보와 테이블 셀 위치를 복사해서 변수로 선언
URL_PAGE1 = "https://www.jeju.go.kr/news/news/law/jeju2.htm"
URL_PAGE2 = "https://www.jeju.go.kr/news/news/law/jeju2.htm"
SELECTOR_PAGE1 = "#app > div.paging-container > div > ul > li:nth-child(1) > a"
SELECTOR_PAGE2 = "#app > div.paging-container > div > ul > li:nth-child(2) > a"
XPATH_PAGE1 = '//*[@id="app"]/div[4]/div/ul/li[1]/a'
XPATH_PAGE2 = '//*[@id="app"]/div[4]/div/ul/li[2]/a'
# 테이블의 첫번째 셀의 정보를 확인
TABLE_XPATH = '//*[@id="gosiBody"]/tr[1]/td[1]'
TABLE_SELECTOR = '#gosiBody > tr:nth-child(1) > td.no'

EXCEL_FILE_NAME = '제주도 기관 공고.xlsx'
EXCEL_SHEET_NAME = '제주도청'
PAGE_COUNT = 5  # 추출할 페이지 수
CELL_FROM = 'B5'  # 데이터 시작 위치

excel_output_path = os.path.join('.', EXCEL_FILE_NAME)
# =======================

def make_template(str1, str2):
    # 가장 긴 공통 접두사
    prefix = ''
    for a, b in zip(str1, str2):
        if a == b:
            prefix += a
        else:
            break
    # 가장 긴 공통 접미사
    rev1, rev2 = str1[::-1], str2[::-1]
    suffix = ''
    for a, b in zip(rev1, rev2):
        if a == b:
            suffix = a + suffix
        else:
            break
    # 템플릿 생성
    template = prefix + '{page}' + suffix
    return template

def extract_table_root_xpath(xpath):
    # '//*[@id="gosiBody"]/tr[1]/td[1]' -> '//*[@id="gosiBody"]'
    return re.sub(r'/tr\[.*?\]/td\[.*?\]$', '', xpath)

def extract_table_root_selector(selector):
    # '#gosiBody > tr:nth-child(1) > td.no' -> '#gosiBody'
    return selector.split('>')[0].strip()

def is_url_template(url1, url2):
    return url1 != url2 and '{page}' in make_template(url1, url2)

def is_xpath_template(xpath1, xpath2):
    return xpath1 != xpath2 and '{page}' in make_template(xpath1, xpath2)

def is_selector_template(sel1, sel2):
    return sel1 != sel2 and '{page}' in make_template(sel1, sel2)

# 템플릿 자동 추출
url_template = make_template(URL_PAGE1, URL_PAGE2)
selector_template = make_template(SELECTOR_PAGE1, SELECTOR_PAGE2)
xpath_template = make_template(XPATH_PAGE1, XPATH_PAGE2)

def crawl_site(page_count, progress_callback, stop_flag):
    service = ChromeService(ChromeDriverManager().install())
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    driver = webdriver.Chrome(service=service, options=options)
    data = []
    try:
        for page in range(1, page_count + 1):
            if stop_flag['stop']:
                progress_callback("중지되었습니다.")
                break
            # 1. URL이 다르면 URL 템플릿 우선
            if is_url_template(URL_PAGE1, URL_PAGE2):
                url = url_template.format(page=page)
                progress_callback(f"{page}페이지 데이터 추출 중입니다... {url}")
                driver.get(url)
                time.sleep(2)
            # 2. XPATH가 다르면 XPATH 템플릿으로 페이지 버튼 클릭
            elif is_xpath_template(XPATH_PAGE1, XPATH_PAGE2):
                if page == 1:
                    driver.get(URL_PAGE1)
                    time.sleep(2)
                else:
                    xpath = xpath_template.format(page=page)
                    try:
                        page_btn = driver.find_element(By.XPATH, xpath)
                        driver.execute_script("arguments[0].scrollIntoView(true);", page_btn)
                        page_btn.click()
                        time.sleep(2)
                    except Exception as e:
                        progress_callback(f"{page}페이지 XPATH 버튼 클릭 실패: {e}")
                        break
            # 3. SELECTOR가 다르면 SELECTOR 템플릿으로 페이지 버튼 클릭
            elif is_selector_template(SELECTOR_PAGE1, SELECTOR_PAGE2):
                if page == 1:
                    driver.get(URL_PAGE1)
                    time.sleep(2)
                else:
                    selector = selector_template.format(page=page)
                    try:
                        page_btn = driver.find_element(By.CSS_SELECTOR, selector)
                        driver.execute_script("arguments[0].scrollIntoView(true);", page_btn)
                        page_btn.click()
                        time.sleep(2)
                    except Exception as e:
                        progress_callback(f"{page}페이지 SELECTOR 버튼 클릭 실패: {e}")
                        break
            else:
                progress_callback("페이지 전환 규칙을 찾을 수 없습니다.")
                break

            # 테이블 루트 자동 추출
            if TABLE_XPATH:
                table_root = extract_table_root_xpath(TABLE_XPATH)
                table_element = driver.find_element(By.XPATH, table_root)
            else:
                table_root = extract_table_root_selector(TABLE_SELECTOR)
                table_element = driver.find_element(By.CSS_SELECTOR, table_root)

            rows = table_element.find_elements(By.TAG_NAME, 'tr')
            for row in rows:
                cols = row.find_elements(By.TAG_NAME, 'td')
                if cols:
                    row_data = [col.text.strip() for col in cols]
                    # 링크 자동 추출 (행 전체에서 a 태그가 있는 첫 번째 셀의 href)
                    link = ""
                    for col in cols:
                        try:
                            a_tag = col.find_element(By.TAG_NAME, "a")
                            link = a_tag.get_attribute("href")
                            break
                        except Exception:
                            continue
                    row_data.append(link)
                    data.append(row_data)
        else:
            progress_callback("완료 되었습니다.")
    except Exception as e:
        progress_callback(f'데이터 추출 중 오류: {e}')
    finally:
        driver.quit()

    # 엑셀 저장 (기존 코드와 동일)
    if data:
        if os.path.exists(excel_output_path):
            wb = openpyxl.load_workbook(excel_output_path)
            if EXCEL_SHEET_NAME in wb.sheetnames:
                ws = wb[EXCEL_SHEET_NAME]
                for row in range(5, ws.max_row + 1):
                    for col in range(2, ws.max_column + 1):
                        ws.cell(row=row, column=col).value = None
            else:
                ws = wb.create_sheet(EXCEL_SHEET_NAME)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = EXCEL_SHEET_NAME

        # B5와 같은 셀 주소를 숫자로 변환
        col_letter, start_row = coordinate_from_string(CELL_FROM)
        start_col = column_index_from_string(col_letter)

        # 데이터 입력
        for i, row in enumerate(data):
            for j, value in enumerate(row):
                ws.cell(row=start_row + i, column=start_col + j, value=value)
            ws.row_dimensions[start_row + i].height = 30

        # 테두리 적용
        for row in range(start_row, ws.max_row + 1):
            for col in range(start_col, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

        wb.save(excel_output_path)
        progress_callback(f"데이터 업데이트 완료: {excel_output_path}")
    else:
        progress_callback("추출된 데이터가 없습니다.")

# --- GUI 구성 ---
class CrawlerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("제주도청 고시공고 추출기")
        self.root.geometry("400x220")  # 창 크기(width x height) 지정
        self.stop_flag = {'stop': False}
        self.thread = None

        self.label = tk.Label(root, text="제주도청 사이트 고시공고를 추출합니다.", font=("맑은 고딕", 12))
        self.label.pack(pady=20, anchor="center")  # 가운데 정렬

        self.progress = tk.Label(root, text="", font=("맑은 고딕", 11))
        self.progress.pack(pady=10, anchor="center")  # 가운데 정렬

        btn_frame = tk.Frame(root)
        btn_frame.pack(pady=10, anchor="center")  # 가운데 정렬

        self.start_btn = tk.Button(btn_frame, text="시작", width=10, command=self.start_crawling)
        self.start_btn.pack(side=tk.LEFT, padx=5)

        self.stop_btn = tk.Button(btn_frame, text="중지", width=10, command=self.stop_crawling, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=5)

        self.exit_btn = tk.Button(btn_frame, text="종료", width=10, command=self.exit_app)
        self.exit_btn.pack(side=tk.LEFT, padx=5)

    def update_progress(self, msg):
        self.progress.config(text=msg)
        self.root.update_idletasks()

    def start_crawling(self):
        self.stop_flag['stop'] = False
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.progress.config(text="크롤링 시작...")
        self.thread = threading.Thread(target=crawl_site, args=(PAGE_COUNT, self.update_progress, self.stop_flag))
        self.thread.start()

    def stop_crawling(self):
        self.stop_flag['stop'] = True
        self.stop_btn.config(state=tk.DISABLED)
        self.start_btn.config(state=tk.NORMAL)

    def exit_app(self):
        if self.thread and self.thread.is_alive():
            self.stop_flag['stop'] = True
            self.thread.join()
        self.root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = CrawlerGUI(root)
    root.mainloop()