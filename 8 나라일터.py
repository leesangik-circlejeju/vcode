import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment

# 설정 변수
CONFIG = {
    'URL': 'https://www.gojobs.go.kr/apmList.do?menuNo=401&mngrMenuYn=N&selMenuNo=400&upperMenuNo=&wd=1207',
    'EXCEL_FILE_NAME': '제주도 기관 공고.xlsx',
    'SHEET_NAME': '나라일터',
    'MAX_PAGES': 3,  # 수집할 최대 페이지 수
    'START_CELL': 'B5',  # 시작 셀
    'ROW_HEIGHT': 25,    # 행 높이
}

COLUMNS = ['번호', '기관유형', '공고명', '기관명', '게시일', '공고마감일', '조회수']

def save_to_excel(data, filename):
    try:
        # 데이터프레임 생성 (헤더 없이)
        df = pd.DataFrame(data, columns=COLUMNS)
        
        # 기존 파일이 있는지 확인
        try:
            wb = load_workbook(filename)
            ws = wb[CONFIG['SHEET_NAME']]
            sheet_exists = True
        except:
            sheet_exists = False
        
        if sheet_exists:
            # 기존 시트에 데이터만 업데이트
            # 셀 주소에서 행과 열 추출 (예: 'B5', 'C10' 등에 대응)
            cell_address = CONFIG['START_CELL']
            col_part = ''.join(filter(str.isalpha, cell_address))  # 알파벳 부분 추출
            row_part = ''.join(filter(str.isdigit, cell_address))  # 숫자 부분 추출
            
            cell_row = start_row = int(row_part)  # 숫자를 정수로 변환
            cell_col = start_col = sum((ord(c.upper()) - ord('A') + 1) * (26 ** i) 
                          for i, c in enumerate(reversed(col_part))) # 열 문자를 숫자로 변환
            
            # 기존 데이터 지우기 (B5부터 시작하는 데이터 영역)
            for row in range(start_row, ws.max_row + 1):
                for col in range(start_col, start_col + len(COLUMNS)):
                    ws.cell(row=row, column=col).value = None
            
            # 새로운 데이터 입력
            for idx, row in df.iterrows():
                cell_row = start_row + idx
                for col_idx, value in enumerate(row):
                    cell_col = start_col + col_idx
                    cell = ws.cell(row=cell_row, column=cell_col)
                    cell.value = value
            wb.save(filename)
        else:
            # 새 파일 생성 (헤더 없이)
            df.to_excel(filename, sheet_name=CONFIG['SHEET_NAME'], index=False, header=False, startrow=4, startcol=1)
        
        print(f"데이터가 {CONFIG['EXCEL_FILE_NAME']} 파일로 저장되었습니다.")
    except Exception as e:
        print(f"엑셀 저장 중 오류 발생: {e}")

def setup_driver():
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--window-size=1920,1080')
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)

def select_jeju_region(driver):
    try:
        # 더보기 버튼 클릭
        more_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "#btnSetSearchView"))
        )
        more_button.click()
        time.sleep(2)

        # 지역 선택 드롭다운 찾기
        region_select = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#serachAreaClassCd"))
        )
        
        # Select 객체 생성
        select = Select(region_select)
        
        # 제주 지역 선택
        select.select_by_value("50000")
        time.sleep(2)
        
        # 검색 버튼 클릭
        search_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "#btnSearch"))
        )
        search_button.click()
        time.sleep(3)
        
        return True
    except Exception as e:
        print(f"지역 선택 중 오류 발생: {e}")
        return False

def go_to_page(driver, page_number):
    try:
        page_link = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, f"#paging > ul > li:nth-child({page_number + 2}) > a"))
        )
        page_link.click()
        time.sleep(2)
        return True
    except Exception as e:
        print(f"페이지 {page_number} 이동 중 오류 발생: {e}")
        return False

def extract_data():
    driver = setup_driver()
    all_data = []
    try:
        driver.get(CONFIG['URL'])
        time.sleep(3)
        
        if not select_jeju_region(driver):
            return
        
        for page in range(1, CONFIG['MAX_PAGES'] + 1):
            print(f"페이지 {page} 처리 중...")
            
            if page > 1:
                if not go_to_page(driver, page):
                    break
            
            rows = driver.find_elements(By.CSS_SELECTOR, "#apmTbl > tbody > tr")
            if len(rows) == 0:
                rows = driver.find_elements(By.CSS_SELECTOR, "table.tbl_list tbody tr")
            
            for row in rows:
                try:
                    tds = row.find_elements(By.TAG_NAME, 'td')
                    if len(tds) < 6:
                        continue
                    
                    try:
                        img = tds[1].find_element(By.TAG_NAME, 'img')
                        COLUMN2 = img.get_attribute('alt')
                    except:
                        COLUMN2 = ''
                    
                    # 웹 페이지의 각 행(row)에서 데이터를 추출
                    COLUMN1 = tds[0].text.strip()  # 공고 번호
                    COLUMN3 = tds[1].text.strip()  # 채용 공고의 제목
                    COLUMN4 = tds[2].text.strip()  # 채용 기관명
                    COLUMN5 = tds[3].text.strip()  # 공고가 게시된 날짜
                    COLUMN6 = tds[4].text.strip()  # 공고 마감 날짜
                    COLUMN7 = tds[5].text.strip()  # 공고 조회수
                    
                    # COLUMNS 리스트의 순서대로 데이터를 딕셔너리에 추가
                    row_data = {}
                    for i, col_name in enumerate(COLUMNS):
                        row_data[col_name] = locals()[f'COLUMN{i+1}']
                    
                    all_data.append(row_data)
                except Exception as e:
                    continue
            
        if all_data:
            save_to_excel(all_data, CONFIG['EXCEL_FILE_NAME'])
        else:
            print("수집된 데이터가 없습니다.")
    except Exception as e:
        print(f"전체 처리 중 오류 발생: {e}")
    finally:
        driver.quit()

if __name__ == "__main__":
    extract_data() 